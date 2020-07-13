import pandas
from statsmodels.robust.scale import huber, Huber
import statsmodels.robust.norms as norms
from openpyxl import load_workbook
import os


def getLocation(df, name):
    loc = df[name].mean()
    try:
        loc, scale = Huber(norm=norms.TukeyBiweight(c=4.685))(df[name])
    except ValueError:
        print('error ' + name)
    return loc
# name = 'Schwank Seminar Ausgewählte Kapitel der Didaktik der Mathematik'
df = pandas.read_excel('ergebnisse.xlsx', sheet_name=0) # can also index sheet by name or fetch all sheets
doneNames = df['Name'].tolist()
for filename in os.listdir():
    if filename.endswith('.xlsx') and not filename.startswith('ergebnis'):
        name = filename.split('.')[0]
        if not name in doneNames:
            print(name)
            dauer = float(input("Dauer: "))
            anzahlAusfaelleVorlesung = float(input("Anzahl Ausfälle: "))
            # anzahlAusfaelleVorlesung = 1
            # dauer = 4.5
            # dauer = 3
            # dauer = 1.5
            kontaktZeitWoche = ((15-anzahlAusfaelleVorlesung) * dauer) * 7/117
            df = pandas.read_excel(open(name + '.xlsx', 'rb'), sheet_name=0)
            df.drop(['Nachname', 'Vorname', 'Benutzername'], axis=1, inplace=True)
            df.columns = ['Bearbeitungsdauer', 'Beendet', 'Lehramt', 'Lehramt1', 'Semester der Veranstaltung', 'Semester', 'Veranstaltungstyp', 'Veranstaltungstyp1',
                        'Anwesenheit1', 'Anwesenheit', 'Woche', 'Klausur', 'Workload 3', 'Angemessen', 'Überschneidungen', 'Überschneidungen1', 'Überschneidungen (2)']
            df.drop(['Lehramt', 'Semester der Veranstaltung', 'Veranstaltungstyp',
                    'Anwesenheit1', 'Workload 3', 'Überschneidungen'], axis=1, inplace=True)
            df.drop(df[df.Woche > 168].index, inplace=True)
            df.drop(df[df.Beendet == '-'].index, inplace=True)
            df.drop(df[df.Semester != 2].index, inplace=True)
            kontaktZeit = []
            selbstZeit = []
            for index, row in df.iterrows():
                anwesenheit = row['Anwesenheit']*0.25
                kZeit = min(max(anwesenheit*kontaktZeitWoche, 0), row['Woche'])
                kontaktZeit.append(kZeit*117/7)
                selbstZeit.append(max((row['Woche'] - kZeit), 0) * 117/7)
            df['Kontaktzeit'] = kontaktZeit
            df['Selbstzeit'] = selbstZeit
            df['Gesamtarbeit'] = df['Selbstzeit'] + df['Kontaktzeit'] + df['Klausur']

            # df['ArbeitRang'] = df['Gesamtarbeit'].rank(ascending = False)
            # df['AngemessenRang'] = df['Angemessen'].rank(ascending = False)
            # df.corr(method = "kendall")
            corr1 = df['Angemessen'].corr(df['Gesamtarbeit'], method="spearman")
            corr2 = df['Angemessen'].corr(df['Klausur'], method="spearman")
            corr3 = df['Angemessen'].corr(df['Selbstzeit'], method="spearman")
            loc = getLocation(df, 'Kontaktzeit')
            loc0 = getLocation(df, 'Angemessen')
            loc1 = getLocation(df, 'Klausur')
            loc2 = getLocation(df, 'Selbstzeit')
            loc3 = getLocation(df, 'Gesamtarbeit')
            dfe = pandas.DataFrame(data={'Name': [name],
                                        'Angemessenheit/Gesamtarbeit': [corr1],
                                        'Angemessenheit/Klausurarbeit': [corr2],
                                        'Angemessenheit/Selbstarbeitszeit': [corr3],
                                        'Angemessenheit': [100-(loc0-1)*100/3],
                                        'Kontaktzeit': [loc],
                                        'Klausurarbeit': [loc1],
                                        'Selbstarbeit': [loc2],
                                        'Gesamtarbeitszeit': [loc3]})
            # with open(name + '.txt', 'w') as the_file:
            #     the_file.write(
            #         'Korrelation Angemessenheit/Gesamtarbeit = ' + str(corr1) + '\n')
            #     the_file.write(
            #         'Korrelation Angemessenheit/Klausurarbeit = ' + str(corr2) + '\n')
            #     the_file.write(
            #         'Korrelation Angemessenheit/Selbstarbeitszeit = ' + str(corr3) + '\n')
            #     the_file.write('Zentrale Tendenz Angemessenheit = ' +
            #                    str(100-(loc0-1)*100/3) + '\n')
            #     the_file.write('Zentrale Tendenz Kontaktzeit = ' + str(loc) + '\n')
            #     the_file.write('Zentrale Tendenz Klausurarbeit = ' + str(loc1) + '\n')
            #     the_file.write('Zentrale Tendenz Selbstarbeit = ' + str(loc2) + '\n')
            #     the_file.write('Zentrale Tendenz Gesamtarbeitszeit = ' + str(loc3) + '\n')
            # writer = pandas.ExcelWriter(name + ' Data.xlsx')
            # df.to_excel(writer, "Main")
            # writer.save()

            book = load_workbook('ergebnisse.xlsx')
            writer = pandas.ExcelWriter('ergebnisse.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            for sheetname in writer.sheets:
                dfe.to_excel(writer, sheet_name=sheetname,
                            startrow=writer.sheets[sheetname].max_row, index=False, header=False)

            writer.save()
            print('Done ' + name)
